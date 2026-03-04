#!/usr/bin/env python3
"""
Scan Fallen Sword guilds A–Z and produce an Excel report of feasible GvG conflicts
given a list of attackers (participants) and their levels.

AUTH:
  - Prefer --cookie (paste browser Cookie header) for reliability.
  - --email/--password-env uses a best-effort form-post login.

DISCLAIMER:
  Respect the game's Terms of Use and rate-limit your requests.
"""

from __future__ import annotations
import sys
from tqdm.auto import tqdm
import argparse
import math
import os
import re
import time
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup

# We use openpyxl objects directly; import lazily to keep startup clean
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation


# -----------------------------
# Constants / URLs
# -----------------------------
FALLENSWORD_BASE = "https://www.fallensword.com/"
GUILD_ATOZ_URL = "https://www.fallensword.com/index.php?cmd=guild&subcmd=atoz&letter={letter}&page={page}"
GUILD_VIEW_URL = "https://www.fallensword.com/index.php?cmd=guild&subcmd=view&guild_id={guild_id}"

HUNTEDCOW_LOGIN_URL = "https://account.huntedcow.com/auth?game=6"

DEFAULT_USER_AGENT = "Mozilla/5.0 (compatible; fs-gvg-scanner/1.0; +https://www.fallensword.com/)"


# -----------------------------
# Data Models
# -----------------------------
@dataclass(frozen=True)
class Guild:
    guild_id: int
    name: str
    members_total: int
    url: str


@dataclass(frozen=True)
class Member:
    player_id: int
    name: str
    level: int
    inactive_days: Optional[int]  # None if unknown
    profile_url: str


@dataclass(frozen=True)
class Attacker:
    name: str
    level: int


# -----------------------------
# Range Rules (from user notes)
# -----------------------------
def range_delta(low_level: int) -> int:
    """Level range delta based on *lowest* player's level (your table)."""
    if low_level < 50:
        return 0
    if 50 <= low_level <= 300:
        return 25
    if 301 <= low_level <= 700:
        return 50
    if 701 <= low_level <= 1000:
        return 100
    # 1001-2000 -> 125, 2001-3000 -> 150, ... +25 per 1000 above 1000
    # Using ceil so 1001..2000 => 1, 2001..3000 => 2, etc.
    steps = math.ceil((low_level - 1000) / 1000)
    return 100 + 25 * steps


def can_attack(attacker_level: int, defender_level: int) -> bool:
    """Return True if attacker can hit defender under the 'lowest level determines delta' rule."""
    if attacker_level < 50 or defender_level < 50:
        return False
    low = min(attacker_level, defender_level)
    delta = range_delta(low)
    return abs(attacker_level - defender_level) <= delta


# -----------------------------
# Parsing helpers
# -----------------------------
_LAST_ACTIVITY_RE = re.compile(r"Last Activity:\s*</td>\s*<td>\s*(\d+)d\b", re.IGNORECASE)

def parse_inactive_days_from_data_tipped(data_tipped: str) -> Optional[int]:
    """
    data-tipped example contains HTML like: ... Last Activity:</td><td>227d 20h 56m 47s</td> ...
    Return days as int, or None if missing.
    """
    if not data_tipped:
        return None
    m = _LAST_ACTIVITY_RE.search(data_tipped)
    if not m:
        return None
    return int(m.group(1))


def safe_int(s: str, default: int = 0) -> int:
    try:
        return int(str(s).strip())
    except Exception:
        return default


# -----------------------------
# HTTP client
# -----------------------------
class FSClient:
    def __init__(
        self,
        *,
        cookie_header: Optional[str],
        email: Optional[str],
        password: Optional[str],
        user_agent: str = DEFAULT_USER_AGENT,
        timeout: int = 30,
        min_delay_s: float = 0.35,
    ):
        self.sess = requests.Session()
        self.sess.headers.update({"User-Agent": user_agent})
        self.timeout = timeout
        self.min_delay_s = min_delay_s
        self._last_request_ts = 0.0

        if cookie_header:
            # Attach cookie header as session cookies (best effort)
            # Cookie header format: "a=b; c=d; ..."
            for part in cookie_header.split(";"):
                part = part.strip()
                if not part or "=" not in part:
                    continue
                k, v = part.split("=", 1)
                self.sess.cookies.set(k.strip(), v.strip())

        self.email = email
        self.password = password

    def _sleep_if_needed(self):
        now = time.time()
        elapsed = now - self._last_request_ts
        if elapsed < self.min_delay_s:
            time.sleep(self.min_delay_s - elapsed)
        self._last_request_ts = time.time()

    def get(self, url: str) -> requests.Response:
        self._sleep_if_needed()
        resp = self.sess.get(url, timeout=self.timeout, allow_redirects=True)
        resp.raise_for_status()
        return resp

    def post(self, url: str, data: dict) -> requests.Response:
        self._sleep_if_needed()
        resp = self.sess.post(url, data=data, timeout=self.timeout, allow_redirects=True)
        resp.raise_for_status()
        return resp

    def is_logged_in(self) -> bool:
        """
        Best-effort check: try to load ATOZ and see if it looks like the public marketing page.
        """
        test_url = GUILD_ATOZ_URL.format(letter="A", page=0)
        html = self.get(test_url).text
        # If not logged in, user usually gets the marketing landing page with "Create a Free Account"
        return "Create a Free Account" not in html and "Already have an account" not in html

    def login_best_effort(self) -> bool:
        """
        Attempt to login via account.huntedcow.com form scraping.

        If the login page is JS-driven or protected, this might fail.
        In that case: use --cookie.
        """
        if not (self.email and self.password):
            return False

        # If already logged in via cookie, keep it.
        if self.is_logged_in():
            return True

        # 1) Load login page
        r = self.get(HUNTEDCOW_LOGIN_URL)
        soup = BeautifulSoup(r.text, "lxml")
        form = soup.find("form")
        if not form:
            return False

        action = form.get("action") or HUNTEDCOW_LOGIN_URL
        method = (form.get("method") or "post").lower()
        post_url = urljoin(HUNTEDCOW_LOGIN_URL, action)

        # 2) Collect inputs
        payload: Dict[str, str] = {}
        email_field = None
        password_field = None

        for inp in form.find_all("input"):
            name = inp.get("name")
            if not name:
                continue
            value = inp.get("value", "")
            payload[name] = value

            t = (inp.get("type") or "").lower()
            if t == "email":
                email_field = name
            if t == "password":
                password_field = name

        # Common fallbacks if types aren’t present
        if email_field is None:
            for k in payload.keys():
                if k.lower() in ("email", "username", "login", "user"):
                    email_field = k
                    break
        if password_field is None:
            for k in payload.keys():
                if "pass" in k.lower():
                    password_field = k
                    break

        if not email_field or not password_field:
            return False

        payload[email_field] = self.email
        payload[password_field] = self.password

        # 3) Submit
        if method == "get":
            _ = self.get(post_url)  # not typical for login
        else:
            _ = self.post(post_url, payload)

        # 4) Verify by trying to access guild pages
        return self.is_logged_in()


# -----------------------------
# Scrapers
# -----------------------------
def parse_guild_list(html: str) -> List[Guild]:
    """
    Parse A–Z guild list page:
      - guild_id from href "guild_id=52645"
      - name from anchor text
      - members from the 'Members' column
    """
    soup = BeautifulSoup(html, "lxml")
    guilds: List[Guild] = []

    # Heuristic: guild rows contain link with cmd=guild&subcmd=view&guild_id=...
    for a in soup.select('a[href*="cmd=guild"][href*="subcmd=view"][href*="guild_id="]'):
        href = a.get("href", "")
        m = re.search(r"guild_id=(\d+)", href)
        if not m:
            continue
        guild_id = int(m.group(1))
        name = a.get_text(strip=True)

        # members count is usually in the next <td> in that row
        tr = a.find_parent("tr")
        if not tr:
            continue
        tds = tr.find_all("td")
        members_total = 0
        if len(tds) >= 3:
            members_total = safe_int(tds[2].get_text(strip=True), 0)

        url = urljoin(FALLENSWORD_BASE, href)
        guilds.append(Guild(guild_id=guild_id, name=name, members_total=members_total, url=url))

    # Deduplicate by id (some pages can duplicate anchors)
    dedup: Dict[int, Guild] = {}
    for g in guilds:
        dedup[g.guild_id] = g
    return list(dedup.values())


def parse_guild_members(html: str) -> List[Member]:
    """
    Parse guild view page members table. We use the anchor with cmd=profile&player_id=...
    and read:
      - player name
      - player_id
      - level (the adjacent <td align="center">)
      - inactive days from data-tipped "Last Activity"
    """
    soup = BeautifulSoup(html, "lxml")
    members: List[Member] = []

    for a in soup.select('a[href*="cmd=profile"][href*="player_id="]'):
        href = a.get("href", "")
        m = re.search(r"player_id=(\d+)", href)
        if not m:
            continue
        player_id = int(m.group(1))
        name = a.get_text(strip=True)

        tr = a.find_parent("tr")
        if not tr:
            continue
        tds = tr.find_all("td")
        # Expected: [status_icon, username_cell, level_cell, rank_cell, contrib_cell]
        level = None
        if len(tds) >= 3:
            level = safe_int(tds[2].get_text(strip=True), 0)

        data_tipped = a.get("data-tipped", "") or ""
        inactive_days = parse_inactive_days_from_data_tipped(data_tipped)

        profile_url = urljoin(FALLENSWORD_BASE, href)
        members.append(
            Member(
                player_id=player_id,
                name=name,
                level=int(level or 0),
                inactive_days=inactive_days,
                profile_url=profile_url,
            )
        )

    # Deduplicate by player_id
    dedup: Dict[int, Member] = {}
    for mbr in members:
        dedup[mbr.player_id] = mbr
    return list(dedup.values())


def iter_letters(letters: Optional[str]) -> List[str]:
    if not letters:
        return [chr(c) for c in range(ord("A"), ord("Z") + 1)]
    parts = [p.strip().upper() for p in letters.split(",") if p.strip()]
    # Allow ranges like A-D
    out: List[str] = []
    for p in parts:
        if "-" in p and len(p) == 3:
            start, end = p.split("-")
            for c in range(ord(start), ord(end) + 1):
                out.append(chr(c))
        else:
            out.append(p)
    return sorted(set(out))


# -----------------------------
# Conflict evaluation
# -----------------------------
def evaluate_guild_conflicts(
    guild: Guild,
    members: List[Member],
    attackers: List[Attacker],
    *,
    active_days_threshold: int = 7,
    min_initiator_participants: int = 3,  # your note: ">2"
    attack_sizes: Tuple[int, ...] = (50, 75, 100),
) -> Tuple[dict, List[dict], Optional[List[dict]]]:
    """
    Returns:
      - summary row dict (for Conflicts sheet)
      - coverage rows list (for AttackerCoverage sheet)
      - optional targets rows list (for Targets sheet) – created elsewhere when requested
    """
    active = [
        m for m in members
        if (m.inactive_days is not None and m.inactive_days < active_days_threshold)
    ]
    active_50plus = [m for m in active if m.level >= 50]

    # For each attacker, count eligible targets
    attacker_target_counts: Dict[str, int] = {}
    attacker_has_any: Dict[str, bool] = {}

    # Precompute for optional Targets sheet too
    target_hitters: Dict[int, List[str]] = {m.player_id: [] for m in active_50plus}

    for a in attackers:
        cnt = 0
        for t in active_50plus:
            if can_attack(a.level, t.level):
                cnt += 1
                target_hitters[t.player_id].append(a.name)
        attacker_target_counts[a.name] = cnt
        attacker_has_any[a.name] = cnt > 0

    viable_attackers = [a for a in attackers if attacker_has_any.get(a.name, False)]
    viable_count = len(viable_attackers)
    cap_attacks = viable_count * 25

    # Decide feasible sizes
    can_size: Dict[int, bool] = {}
    for sz in attack_sizes:
        required_by_cap = math.ceil(sz / 25)  # each attacker max 25 attacks
        required = max(required_by_cap, min_initiator_participants)
        can_size[sz] = viable_count >= required

    recommended = None
    for sz in sorted(attack_sizes):
        if can_size.get(sz):
            recommended = sz
    # recommended is the largest feasible (since iter sorted asc, last True wins)
    # If none feasible, stays None.

    min_lvl = min([m.level for m in active_50plus], default=None)
    max_lvl = max([m.level for m in active_50plus], default=None)

    summary = {
        "guild_id": guild.guild_id,
        "guild_name": guild.name,
        "members_total": guild.members_total,
        "active_members_<7d": len(active),
        "active_50plus_<7d": len(active_50plus),
        "min_lvl_active": min_lvl,
        "max_lvl_active": max_lvl,
        "viable_attackers": viable_count,
        "cap_attacks": cap_attacks,
        "can_50": can_size.get(50, False),
        "can_75": can_size.get(75, False),
        "can_100": can_size.get(100, False),
        "recommended": recommended or "",
        "guild_url": guild.url,
    }

    coverage_rows = []
    for a in attackers:
        coverage_rows.append({
            "guild_id": guild.guild_id,
            "guild_name": guild.name,
            "attacker": a.name,
            "attacker_level": a.level,
            "eligible_targets_active_<7d": attacker_target_counts.get(a.name, 0),
        })

    return summary, coverage_rows, None


# -----------------------------
# Excel formatting
# -----------------------------
def autosize_worksheet_columns(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, max_width: int = 60):
    ws = writer.sheets[sheet_name]
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].head(500).tolist()])
        width = min(max_len + 2, max_width)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

def format_guild_attacker_matrix_sheet(ws, df: pd.DataFrame):
    """
    Apply:
    - freeze panes / header style
    - dropdown validation with only ✓ or X
    - green/red conditional formatting
    """
    beautify_sheet(ws)
    autosize_openpyxl(ws, df)

    if df.empty or df.shape[1] <= 1:
        return

    max_row = ws.max_row
    max_col = ws.max_column

    # Matrix area excludes first column (guild_name)
    start_row = 2
    start_col = 2
    end_row = max_row
    end_col = max_col

    matrix_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    # Center align matrix cells
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Wider first column for guild names
    ws.column_dimensions["A"].width = min(max(ws.column_dimensions["A"].width, 32), 60)

    # Dropdown: only ✓ or X
    dv = DataValidation(
        type="list",
        formula1='"✓,X"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.prompt = "Choose ✓ or X"
    dv.promptTitle = "Guild/Attacker Flag"
    dv.error = "Only ✓ or X is allowed."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(matrix_range)

    # Conditional formatting
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    red_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

    ws.conditional_formatting.add(
        matrix_range,
        CellIsRule(operator="equal", formula=['"✓"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        matrix_range,
        CellIsRule(operator="equal", formula=['"X"'], fill=red_fill)
    )




def beautify_sheet(ws):
    ws.freeze_panes = "A2"
    # Header style
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.auto_filter.ref = ws.dimensions


def autosize_openpyxl(ws, df: pd.DataFrame, max_width: int = 65):
    for idx, col in enumerate(df.columns, start=1):
        values = [str(col)] + [str(v) for v in df[col].head(500).tolist()]
        width = min(max(len(v) for v in values) + 2, max_width)
        ws.column_dimensions[get_column_letter(idx)].width = width


# -----------------------------
# Main
# -----------------------------
def load_attackers_csv(path: str) -> List[Attacker]:
    df = pd.read_csv(path)
    out: List[Attacker] = []
    for _, r in df.iterrows():
        out.append(Attacker(name=str(r["name"]).strip(), level=int(r["level"])))
    return out

def level_bracket_info(level: int) -> dict:
    """
    Returns the attack bracket info based on the player's own level.
    This is the bracket used when the attacker is the lower/equal side.
    """
    if level < 50:
        return {
            "bracket": "Below 50 (not conflict-eligible)",
            "delta": 0,
            "band_start": None,
            "band_end": None,
        }
    if 50 <= level <= 300:
        return {"bracket": "50-300", "delta": 25, "band_start": 50, "band_end": 300}
    if 301 <= level <= 700:
        return {"bracket": "301-700", "delta": 50, "band_start": 301, "band_end": 700}
    if 701 <= level <= 1000:
        return {"bracket": "701-1000", "delta": 100, "band_start": 701, "band_end": 1000}

    # 1001-2000 -> 125, 2001-3000 -> 150, ...
    band_idx = (level - 1001) // 1000
    band_start = 1001 + band_idx * 1000
    band_end = band_start + 999
    delta = 125 + band_idx * 25

    return {
        "bracket": f"{band_start}-{band_end}",
        "delta": delta,
        "band_start": band_start,
        "band_end": band_end,
    }


def exact_lowest_hittable_level(attacker_level: int) -> Optional[int]:
    """
    Exact lowest target level the attacker can hit under the 'lowest level determines range' rule.
    We brute-force downward because the lower-side delta depends on the target's level.
    """
    if attacker_level < 50:
        return None

    for target_level in range(50, attacker_level + 1):
        if can_attack(attacker_level, target_level):
            return target_level
    return None


def exact_highest_hittable_level(attacker_level: int) -> Optional[int]:
    """
    Exact highest target level the attacker can hit.
    For higher targets, the attacker is the lower level, so the range is based on attacker_level.
    """
    if attacker_level < 50:
        return None
    return attacker_level + range_delta(attacker_level)


def build_guild_attacker_matrix(df_cov: pd.DataFrame, df_conf: pd.DataFrame, attackers: List[Attacker]) -> pd.DataFrame:
    """
    Rows   = guilds
    Cols   = attackers
    Values = ✓ if that attacker has at least one valid active target in that guild, else X

    Only keep rows where at least 2 attackers have valid targets.
    """
    ordered_attackers = [a.name for a in attackers]

    if df_cov.empty:
        return pd.DataFrame(columns=["guild_name"] + ordered_attackers)

    work = df_cov.copy()
    work["guild_label"] = work["guild_name"] + " (" + work["guild_id"].astype(str) + ")"
    work["has_targets"] = work["eligible_targets_active_<7d"] > 0

    pivot = work.pivot_table(
        index="guild_label",
        columns="attacker",
        values="has_targets",
        aggfunc="max",
        fill_value=False,
    )

    # Preserve attacker order from attackers.csv
    pivot = pivot.reindex(columns=ordered_attackers, fill_value=False)

    # Preserve guild order from Conflicts sheet
    if not df_conf.empty:
        guild_order = (df_conf["guild_name"] + " (" + df_conf["guild_id"].astype(str) + ")").tolist()
        guild_order = [g for g in guild_order if g in pivot.index]
        pivot = pivot.reindex(guild_order)

    # Keep only guilds where 2 or more attackers have targets
    pivot = pivot[pivot.sum(axis=1) >= 2]

    # Convert bool -> symbols
    pivot = pivot.astype(bool).replace({True: "✓", False: "X"})

    pivot = pivot.reset_index().rename(columns={"guild_label": "guild_name"})
    return pivot


def build_attacker_sanity_df(attackers: List[Attacker]) -> pd.DataFrame:
    """
    For each attacker, show:
    - own bracket
    - own delta
    - exact lowest hittable target level
    - exact highest hittable target level
    """
    rows = []
    for attacker in attackers:
        info = level_bracket_info(attacker.level)
        lowest = exact_lowest_hittable_level(attacker.level)
        highest = exact_highest_hittable_level(attacker.level)

        rows.append({
            "attacker": attacker.name,
            "attacker_level": attacker.level,
            "own_bracket": info["bracket"],
            "own_delta_for_higher_targets": info["delta"],
            "lowest_hittable_target_level": lowest,
            "highest_hittable_target_level": highest,
            "target_span_width": (highest - lowest) if (lowest is not None and highest is not None) else None,
        })

    return pd.DataFrame(rows)


def main():
    ap = argparse.ArgumentParser()

    ap.add_argument("--attackers", required=True, help="CSV with columns: name,level")
    ap.add_argument("--out", default="conflicts.xlsx", help="Output xlsx path")
    ap.add_argument("--letters", default=None, help="Letters to scan, e.g. A,B,C or A-D. Default: A-Z")
    ap.add_argument("--max-guilds", type=int, default=0, help="Stop after N guilds (0 = no limit)")
    ap.add_argument("--active-days", type=int, default=7, help="Active if inactive_days < this")
    ap.add_argument("--min-participants", type=int, default=2, help="Min attackers required to initiate (over 2 attackers)")
    ap.add_argument("--cookie", default=None, help='Paste browser "Cookie" request header here (recommended)')
    ap.add_argument("--email", default=None, help="Login email (best-effort login)")
    ap.add_argument("--password-env", default=None, help="Env var name holding password (best-effort login)")
    ap.add_argument("--include-targets", action="store_true", help="Add Targets sheet (can get large)")
    ap.add_argument("--min-delay", type=float, default=0.35, help="Delay between requests (rate-limit)")
    ap.add_argument("--cookie-file", default=None, help="Path to a txt file containing the Cookie header value")
    ap.add_argument("--no-progress", action="store_true", help="Disable progress bars")

    args = ap.parse_args()

    progress_disabled = args.no_progress or (not sys.stderr.isatty())

    attackers = load_attackers_csv(args.attackers)
    password = os.getenv(args.password_env) if args.password_env else None

    cookie_header = args.cookie

    if args.cookie_file:
        with open(args.cookie_file, "r", encoding="utf-8") as f:
            cookie_header = f.read().strip()
            # Allow people to paste full line like: Cookie: a=b; c=d
            if cookie_header.lower().startswith("cookie:"):
                cookie_header = cookie_header.split(":", 1)[1].strip()

    client = FSClient(
        cookie_header=cookie_header,
        email=args.email,
        password=password,
        min_delay_s=args.min_delay,
    )

    if not client.is_logged_in():
        ok = client.login_best_effort()
        if not ok:
            raise SystemExit(
                "Not logged in. Use --cookie (recommended) or provide --email and --password-env.\n"
                "Tip: copy the Cookie header from DevTools -> Network -> a request to fallensword.com."
            )

    letters = iter_letters(args.letters)

    summaries: List[dict] = []
    coverages: List[dict] = []
    targets_rows: List[dict] = []

    guild_count = 0

    letters_pbar = tqdm(letters, desc="Scanning letters", unit="letter", disable=progress_disabled)

    for letter in letters_pbar:
        # Inner bar: guilds found for THIS letter (total grows as we discover pages)
        guilds_pbar = tqdm(
            total=0,
            desc=f"Letter {letter} - guilds",
            unit="guild",
            leave=False,
            disable=progress_disabled,
        )

        try:
            for page in range(0, 10):
                letters_pbar.set_postfix_str(f"letter={letter}, page={page+1}")

                url = GUILD_ATOZ_URL.format(letter=letter, page=page)
                html = client.get(url).text
                guilds = parse_guild_list(html)

                if not guilds:
                    break  # no more pages for this letter

                # Increase inner progress total as we discover more guilds
                guilds_pbar.total += len(guilds)
                guilds_pbar.refresh()

                for g in guilds:
                    guild_count += 1
                    if args.max_guilds and guild_count > args.max_guilds:
                        break

                    # Show current guild name in the inner progress bar
                    guilds_pbar.set_postfix_str(f"{g.name} (id={g.guild_id})")

                    g_html = client.get(GUILD_VIEW_URL.format(guild_id=g.guild_id)).text
                    members = parse_guild_members(g_html)

                    summary, coverage_rows, _ = evaluate_guild_conflicts(
                        g,
                        members,
                        attackers,
                        active_days_threshold=args.active_days,
                        min_initiator_participants=args.min_participants,
                    )
                    summaries.append(summary)
                    coverages.extend(coverage_rows)

                    if args.include_targets:
                        active_50plus = [
                            m for m in members
                            if (m.inactive_days is not None and m.inactive_days < args.active_days and m.level >= 50)
                        ]
                        for t in active_50plus:
                            hitters = [a.name for a in attackers if can_attack(a.level, t.level)]
                            if not hitters:
                                continue
                            targets_rows.append({
                                "guild_id": g.guild_id,
                                "guild_name": g.name,
                                "player_id": t.player_id,
                                "player_name": t.name,
                                "level": t.level,
                                "inactive_days": t.inactive_days,
                                "hitters": ", ".join(hitters),
                                "profile_url": t.profile_url,
                            })

                    guilds_pbar.update(1)

                if args.max_guilds and guild_count > args.max_guilds:
                    break

            if args.max_guilds and guild_count > args.max_guilds:
                break

        finally:
            guilds_pbar.close()

    letters_pbar.close()

    # Export
    df_conf = pd.DataFrame(summaries).sort_values(
        by=["recommended", "can_100", "can_75", "can_50", "active_50plus_<7d"],
        ascending=[False, False, False, False, False],
        na_position="last",
    )
    df_cov = pd.DataFrame(coverages)
    df_matrix = build_guild_attacker_matrix(df_cov, df_conf, attackers)
    df_sanity = build_attacker_sanity_df(attackers)

    with pd.ExcelWriter(args.out, engine="openpyxl") as writer:
        # Sheet 1
        df_conf.to_excel(writer, index=False, sheet_name="Conflicts")

        # Sheet 2
        df_cov.to_excel(writer, index=False, sheet_name="AttackerCoverage")

        # Sheet 3
        df_matrix.to_excel(writer, index=False, sheet_name="GuildAttackerMatrix")

        # Sheet 4
        df_sanity.to_excel(writer, index=False, sheet_name="AttackerBracketSanity")

        # Optional Sheet 5
        if args.include_targets:
            pd.DataFrame(targets_rows).to_excel(writer, index=False, sheet_name="Targets")

        for name in writer.sheets:
            ws = writer.sheets[name]

            if name == "Conflicts":
                beautify_sheet(ws)
                autosize_openpyxl(ws, df_conf)

            elif name == "AttackerCoverage":
                beautify_sheet(ws)
                autosize_openpyxl(ws, df_cov)

            elif name == "GuildAttackerMatrix":
                format_guild_attacker_matrix_sheet(ws, df_matrix)

            elif name == "AttackerBracketSanity":
                beautify_sheet(ws)
                autosize_openpyxl(ws, df_sanity)

            elif name == "Targets":
                beautify_sheet(ws)
                autosize_openpyxl(ws, pd.DataFrame(targets_rows))

    print(f"Saved: {args.out}")
    print(f"Guilds scanned: {len(df_conf)}")


if __name__ == "__main__":
    main()